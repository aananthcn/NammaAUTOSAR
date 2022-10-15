#
# Created on Sun Oct 02 2022 10:08:38 AM
#
# The MIT License (MIT)
# Copyright (c) 2022 Aananth C N
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of this software
# and associated documentation files (the "Software"), to deal in the Software without restriction,
# including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so,
# subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all copies or substantial
# portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
# TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
# TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#
import os, sys
import openpyxl

from colorama import Fore, Back, Style
from datetime import datetime


def clear_old_excel_rows(book):
    sheetnames = book.sheetnames
    for sheetname in sheetnames:
        sheet = book[sheetname]
        header_row = 2 # First row will hold disclaimer, title in 2nd row
        active_rows = len(sheet['A'])
        if active_rows < header_row:
            print("Info: sheet \""+sheetname+"\" is new / fresh! So, not clearing old data!")
            return
        sheet.delete_rows(header_row+1, active_rows)        


def open_excel_file(file):
    if os.path.isfile(file):
        try:
            book = openpyxl.load_workbook(file, read_only=False, data_only=True)
        except:
            print(Fore.YELLOW +"Error: Please close " + file + " and retry!")
            print(Style.RESET_ALL)
            return None
    else:
        print("Creating \"" + file + "\"")
        book = openpyxl.Workbook()
    return book


# Function: locate_text_in_column
# -------------------------------
# This function locates the 'text' passed as argument in the excel sheet
# in a "specified" column number passed as argument 'col'
def locate_text_in_column(text, sheet, col):
    if sheet == None or text == None:
        return -1, -1
    if col < 65 or col > 65+26:
        return -1, -1
    
    rows = len(sheet[chr(col)])

    # loop rows and columns to find the title
    for r in range(rows):
        row = r + 1
        if sheet[chr(col)+str(row)].value == None:
            continue
        if text.lower() in str(sheet[chr(col)+str(row)].value).lower():
            return col, row
    print("Info: can't find \'"+ str(text)+"\' by locate_text_in_column()")
    return -1, -1



# Function: locate_text_in_row
# ----------------------------
# This function locates the 'text' passed as argument in the excel sheet
# in a "specified" row number passed as argument 'row'
def locate_text_in_row(text, sheet, row):
    if sheet == None or text == None:
        return -1, -1
    if row < 0 or row > 65535:
        return -1, -1
    
    cols = len(sheet[str(row)])

    # loop rows and columns to find the title
    for c in range(cols):
        col = c + 65
        if sheet[chr(col)+str(row)].value == None:
            continue
        if text.lower() in str(sheet[chr(col)+str(row)].value).lower():
            return col, row
    print("Info: can't find \'"+ str(text)+"\' by locate_text_in_row()")
    return -1, -1



# Function: locate_text_in_cell
# -------------------------------
# This function locates the 'text' passed as argument in the excel sheet
# in a any column or row in the excel sheet
def locate_text_in_cell(text, sheet):
    if sheet == None or text == None:
        return -1, -1
    
    rows = len(sheet['A'])

    # loop rows and columns to find the title
    for r in range(rows):
        row = r + 1
        cols = len(sheet[str(row)])
        for c in range(cols):
            col = c + 65
            #print(sheet[chr(col)+str(row)].value)
            if sheet[chr(col)+str(row)].value == None:
                continue
            if text.lower() in str(sheet[chr(col)+str(row)].value).lower():
                return col, row
    print("Info: can't find \'"+ str(text)+"\' by locate_text_in_cell()")
    return -1, -1



# Function: locate_heading_row
# ----------------------------
# This function locates the heading row in excel sheet. S.No is the 
# key text this function uses to identify it. 
def locate_heading_row(wb, sheetname):
    if sheetname == None or len(sheetname) < 1:
        return -1
    # open sheet and find the number of rows
    sheet = wb[sheetname]
    col, row = locate_text_in_cell("S.No", sheet)
    if col == -1:
        print("Warning: locate_heading_row.locate_text_in_cell returned col == -1")
    if row == -1:
        emsg = "Error: Can't find \"S.No\" in sheet: \"" + sheetname + "\""
        print(emsg)
    
    return row

        

# Function: locate_heading_column
# -------------------------------
# This function locates the heading column in a given excel sheet. Note: S.No 
# will be the key it uses to locate rows, check locate_heading_row()
def locate_heading_column(text, wb, sheetname):
    if wb == None:
        return -1, -1
    if sheetname == None or len(sheetname) < 1 or text == None or len(text) < 1:
        return -1, -1
    
    # locate the row of the heading
    h_row = locate_heading_row(wb, sheetname)
    if h_row != -1:
        # open sheet and locate column in heading row
        sheet = wb[sheetname]
        return locate_text_in_row(text, sheet, h_row)


# Function: excel_int_to_date
# ---------------------------
# Excel converts dates to integer (some times). This function as per stackoverflow converts
# https://stackoverflow.com/questions/31359150/convert-date-from-excel-in-number-format-to-date-format-python
def excel_int_to_date(int_val):
    date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int_val - 2)
    return date


# Function: locate_cols
# ---------------------------
# This function takes a list of column headings as input and find the excel columns
# such as 'A', 'B', 'E", for the correponding column heading strings. 

# Return value: 
# 1) A dictionary of {"column headings" : "[A-Z]"}
# 2) Row in which the header column is placed - An integer.
def locate_cols(wb, sheetname, headers):
    Cols = {}
    Row = None
    for item in headers:
        col, row = locate_heading_column(item, wb, sheetname)
        if col == -1:
            print("Error: cannot locate \""+item+"\" in sheet "+ sheetname)
            return None
        Cols[item] = chr(col)
        if row != -1:
            Row = row

    return Cols, Row
